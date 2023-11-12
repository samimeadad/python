
import openpyxl

dataPath = "D:\\Training\\Big-Data-Data-Analytics-Data-Science-Pondit\\Excel Files\\duplicate-data.xlsx"

df = openpyxl.load_workbook(dataPath)

print(df)

df1 = df.active

for row in range(0, df1.max_row):
    for col in df1.iter_cols(1, df1.max_column):
        print(col[row].value)
